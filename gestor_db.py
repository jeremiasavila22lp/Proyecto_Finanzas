import sqlite3
import psycopg2
import logging
from datetime import datetime
from typing import List, Tuple
import csv
import matplotlib.pyplot as plt
from passlib.context import CryptContext
import threading
from psycopg2.extras import RealDictCursor

# Configuración de Passlib para hashing seguro (Bcrypt)
# Lo definimos a nivel de módulo para que sea accesible por todas las clases
# bcrypt__truncate_error=False permite que passlib trunque automáticamente a 72 bytes sin lanzar error
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto", bcrypt__truncate_error=False)
 

class GestorGastos: 
    def __init__(self, finanzas_bd):
        self.lock = threading.RLock()

        if finanzas_bd.startswith("postgres://") or finanzas_bd.startswith("postgresql://"):
            self.es_postgresql = True
            try:
                self.conexion = psycopg2.connect(finanzas_bd)
                self.cursor = self.conexion.cursor(cursor_factory=RealDictCursor) 
                print("[OK] Conectado a PostgreSQL (Neon)")
            except Exception as e:
                print(f"[ERROR] No se pudo conectar a PostgreSQL: {e}")
                logging.error(f"Error de conexión a la BD: {e}")
                raise e # Re-lanzamos para que se vea en los logs de Render
        else:
            self.es_postgresql = False
            self.conexion = sqlite3.connect(finanzas_bd, check_same_thread=False)
            self.conexion.row_factory = sqlite3.Row
            self.cursor = self.conexion.cursor()
            print("[OK] Conectado a SQLite Local")
        
        self.p = "%s" if self.es_postgresql else "?"
        self.serial_type = "SERIAL PRIMARY KEY" if self.es_postgresql else "INTEGER PRIMARY KEY AUTOINCREMENT"
        self.default_date = "CURRENT_TIMESTAMP" if self.es_postgresql else "CURRENT_TIMESTAMP"
        # Crear tabla de usuarios
        self.cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS usuarios (
                id {self.serial_type},
                nombre TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                fecha_creacion TEXT DEFAULT {self.default_date},
                presupuesto REAL DEFAULT 5000.0
            )
        """)
        # Migración: Verificar si existe la columna 'fecha', si no, agregarla
        if self.es_postgresql:
            self.cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'usuarios'")
            columnas_u = [fila['column_name'] for fila in self.cursor.fetchall()]
        else:
            self.cursor.execute("PRAGMA table_info(usuarios)")
            columnas_u = [col[1] for col in self.cursor.fetchall()]

        if 'presupuesto' not in columnas_u:
            self.cursor.execute("ALTER TABLE usuarios ADD COLUMN presupuesto REAL DEFAULT 5000.0")
            self.conexion.commit()
            print("[OK] Columna 'presupuesto' agregada a la tabla usuarios.")
        else:
            print("[OK] Columna 'presupuesto' ya existe en la tabla usuarios.")
        
        self.cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS gastos (
                id {self.serial_type},
                monto REAL,
                categoria TEXT,
                descripcion TEXT,
                fecha TEXT DEFAULT {self.default_date},
                usuario_id INTEGER,
                FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
            )
        """)
        self.conexion.commit()

        if self.es_postgresql:
            self.cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'gastos'")
            columnas = [fila['column_name'] for fila in self.cursor.fetchall()]
        else:
            self.cursor.execute("PRAGMA table_info(gastos)")
            columnas = [col[1] for col in self.cursor.fetchall()]

        if 'fecha' not in columnas:
            self.cursor.execute("ALTER TABLE gastos ADD COLUMN fecha TEXT")
            # Actualizar registros viejos con fecha actual
            f_now = "CURRENT_DATE" if self.es_postgresql else "date('now')"
            self.cursor.execute(f"UPDATE gastos SET fecha = {f_now} WHERE fecha IS NULL")
            self.conexion.commit()
        else:
            print("[OK] Columna 'fecha' ya existe en la tabla gastos.")
            
        # Migración: Verificar si existe la columna 'usuario_id', si no, agregarla
        if 'usuario_id' not in columnas:
            self.cursor.execute("ALTER TABLE gastos ADD COLUMN usuario_id INTEGER")
            self.conexion.commit()
        else:
            print("[OK] Columna 'usuario_id' ya existe en la tabla gastos.")
        
        self.conexion.commit()



    def agregar_gasto(self, monto, categoria, descripcion, usuario_id=None):
        try:
            # Insertamos con la fecha actual y el usuario_id
            fecha_sql = "CURRENT_DATE" if self.es_postgresql else "date('now','localtime')"
            query = f"INSERT INTO gastos (monto, categoria, descripcion, fecha, usuario_id) VALUES ({self.p}, {self.p}, {self.p}, {fecha_sql}, {self.p})"
            self.cursor.execute(query, (float(monto), categoria.lower(), descripcion, usuario_id))
            self.conexion.commit()
            return True
        except ValueError:
            return False
        except Exception as e:
            logging.error(f"Error al agregar gasto: {e}")
            return False

    def obtener_total(self, usuario_id=None):
        try:
            with self.lock:
                if usuario_id is not None:
                    self.cursor.execute(f"SELECT SUM(monto) FROM gastos WHERE usuario_id = {self.p}", (usuario_id,))
                else:
                    self.cursor.execute("SELECT SUM(monto) FROM gastos")
                resultado = self.cursor.fetchone()
                # El SUM retorna None si la tabla está vacía, lo convertimos a 0
                return float(resultado[0]) if resultado and resultado[0] is not None else 0.0
        except Exception as e:
            logging.error(f"Error critico al obtener el total: {e}")
            return 0.0
    
    def filtrar_por_categoria(self, cat):
        query = f"SELECT monto, categoria, descripcion, fecha FROM gastos WHERE categoria = {self.p}"
        self.cursor.execute(query, (cat.lower(),))
        return self.cursor.fetchall()
    
    def obtencion_del_promedio(self): 
        self.cursor.execute("SELECT AVG(monto) FROM gastos")
        resultado = self.cursor.fetchone()
        return resultado[0] if resultado[0] else 0
    
    def validar_monto(self, monto):
        try:
            valor = float(monto)
            if valor <= 0:
                return None
            return valor
        except ValueError:
            return None
    
    def obtener_todos_los_gastos(self, usuario_id=None):
        """Obtiene todos los gastos de la base de datos, opcionalmente filtrados por usuario"""
        with self.lock:
            if usuario_id is not None:
                self.cursor.execute(f"SELECT id, monto, categoria, fecha, descripcion FROM gastos WHERE usuario_id = {self.p} ORDER BY fecha DESC", (usuario_id,))
            else:
                self.cursor.execute("SELECT id, monto, categoria, fecha, descripcion FROM gastos ORDER BY fecha DESC")
            resultados = self.cursor.fetchall()
            # Convertimos cada fila de SQLite a un Diccionario real de Python
            return [dict(fila) for fila in resultados]

    def obtener_gastos_ordenados_por_monto(self):
        """Obtiene todos los gastos ordenados por monto de mayor a menor"""
        self.cursor.execute("SELECT id, monto, categoria, fecha, descripcion FROM gastos ORDER BY monto DESC")
        resultados = self.cursor.fetchall()
        # Convertimos cada fila de SQLite a un Diccionario real de Python
        return [dict(fila) for fila in resultados]

    def obtener_gastos_por_rango(self, fecha_inicio, fecha_fin):
        """Retorna un diccionario {categoria: total} con gastos en el rango de fechas"""
        query = f"""
            SELECT categoria, SUM(monto) 
            FROM gastos 
            WHERE fecha BETWEEN {self.p} AND {self.p} 
            GROUP BY categoria
        """
        self.cursor.execute(query, (fecha_inicio, fecha_fin))
        resultados = self.cursor.fetchall()
        return {cat: monto for cat, monto in resultados}

    def obtener_todos_los_gastos_filtrados(self, fecha_inicio, fecha_fin):
        """Obtiene los gastos filtrados por fecha como diccionarios"""
        query = f"SELECT id, monto, categoria, fecha, descripcion FROM gastos WHERE fecha BETWEEN {self.p} AND {self.p} ORDER BY fecha DESC"
        self.cursor.execute(query, (fecha_inicio, fecha_fin))
        resultados = self.cursor.fetchall()
        return [dict(fila) for fila in resultados]

    def export_a_csv(self, nombre_archivo = "Reporte_gastos.csv"):
        self.cursor.execute("SELECT * FROM gastos")
        datos = self.cursor.fetchall()

        with open(nombre_archivo, "w", newline="", encoding="utf-8") as archivo:
            escritor = csv.writer(archivo)
            escritor.writerow(["ID", "Monto", "categoria", "Descripcion", "Fecha"])
            escritor.writerows(datos)

        return True

    def eliminar_gasto(self, id_gasto, usuario_id=None):
        """Elimina un gasto específico por su ID"""
        try:
            if usuario_id is not None:
                self.cursor.execute(f"DELETE FROM gastos WHERE id = {self.p} AND usuario_id = {self.p}", (id_gasto, usuario_id))
                # Verificar si se borró algo (si rowcount es 0, es que no existía o no era de ese usuario)
                if self.cursor.rowcount == 0:
                    return False
            else:
                self.cursor.execute(f"DELETE FROM gastos WHERE id = {self.p}", (id_gasto,))
            
            self.conexion.commit()
            return True
        except Exception as e:
            return False

    def actualizar_gasto(self, id_gasto, nuevo_monto, categoria, descripcion, fecha, usuario_id=None):
        """Actualiza la información de un gasto existente"""
        try:
            # Validar que el monto sea número
            monto_float = float(nuevo_monto)
            
            if usuario_id is not None:
                query = f"""
                    UPDATE gastos 
                    SET monto = {self.p}, categoria = {self.p}, descripcion = {self.p}, fecha = {self.p}
                    WHERE id = {self.p} AND usuario_id = {self.p}
                """
                self.cursor.execute(query, (monto_float, categoria, descripcion, fecha, id_gasto, usuario_id))
                if self.cursor.rowcount == 0:
                    return False
            else:
                query = f"""
                    UPDATE gastos 
                    SET monto = {self.p}, categoria = {self.p}, descripcion = {self.p}, fecha = {self.p}
                    WHERE id = {self.p}
                """
                self.cursor.execute(query, (monto_float, categoria, descripcion, fecha, id_gasto))
                
            self.conexion.commit()
            return True
        except ValueError:
            return False
        except Exception as e:
            return False

    def eliminar_todos_los_gastos(self):
        """Elimina todos los registros de la tabla gastos"""
        try:
            self.cursor.execute("DELETE FROM gastos")
            self.conexion.commit()
            return True
        except Exception as e:
            return False

    def buscar_gastos(self, texto):
        """Busca gastos y los devuelve como diccionarios"""
        query = f"""
            SELECT id, monto, categoria, fecha, descripcion 
            FROM gastos
            WHERE descripcion LIKE {self.p} OR categoria LIKE {self.p} OR fecha LIKE {self.p}
            ORDER BY fecha DESC
        """
        param = f"%{texto}%"
        self.cursor.execute(query, (param, param, param))
        resultados = self.cursor.fetchall()
        return [dict(fila) for fila in resultados]

    # ============ MÉTODOS DE AUTENTICACIÓN ============
    
    def registrar_usuario(self, nombre, email, password):
        """
        Registra un nuevo usuario en la base de datos.
        Retorna (True, user_id) si fue exitoso, (False, mensaje_error) si falló
        """
        try:
            # Bcrypt tiene un límite de 72 bytes. Passlib prefiere strings.
            # Si el password es un string muy largo, lo truncamos asegurando que su versión bytes no exceda 72.
            if isinstance(password, bytes):
                password = password.decode('utf-8', errors='ignore')
            
            # Truncado seguro: convertimos a bytes, truncamos a 72 y volvemos a decodificar a string
            pw_bytes = password.encode('utf-8')
            if len(pw_bytes) > 72:
                password = pw_bytes[:72].decode('utf-8', errors='ignore')

            # Hashing seguro con Passlib (pasando el string ya saneado)
            password_hash = pwd_context.hash(password)
            
            if self.es_postgresql:
                self.cursor.execute(f"""
                    INSERT INTO usuarios (nombre, email, password_hash) 
                    VALUES ({self.p}, {self.p}, {self.p}) RETURNING id
                """, (nombre, email, password_hash))
                fila = self.cursor.fetchone()
                user_id = fila['id'] if isinstance(fila, dict) else fila[0]
            else:
                self.cursor.execute(f"""
                    INSERT INTO usuarios (nombre, email, password_hash) 
                    VALUES ({self.p}, {self.p}, {self.p})
                """, (nombre, email, password_hash))
                user_id = self.cursor.lastrowid
            
            self.conexion.commit()
            
            # Obtener el ID del usuario recién creado
            return (True, user_id)
        except (sqlite3.IntegrityError, psycopg2.IntegrityError):
            # Email ya existe
            return (False, "El email ya está registrado")
        except Exception as e:
            logging.error(f"Error al registrar usuario: {e}")
            return (False, f"Error al crear el usuario: {str(e)}")
    
    def verificar_login(self, email, password):
        """
        Verifica las credenciales del usuario.
        Retorna (True, user_data) si es correcto, (False, mensaje_error) si falla
        """
        try:
            # Bcrypt tiene un límite de 72 bytes. Passlib prefiere strings.
            if isinstance(password, bytes):
                password = password.decode('utf-8', errors='ignore')

            # Truncado seguro para verificar (mismo criterio que en el registro)
            pw_bytes = password.encode('utf-8')
            if len(pw_bytes) > 72:
                password = pw_bytes[:72].decode('utf-8', errors='ignore')

            self.cursor.execute(f"""
                SELECT id, nombre, email, password_hash, fecha_creacion, presupuesto 
                FROM usuarios 
                WHERE email = {self.p}
            """, (email,))
            
            resultado = self.cursor.fetchone()
            
            if resultado:
                user_data = dict(resultado)
                stored_hash = user_data.pop('password_hash')
                
                if pwd_context.verify(password, stored_hash):
                    return (True, user_data)
                else:
                    return (False, "Contraseña incorrecta")
            else:
                return (False, "Email no registrado")
        except Exception as e:
            logging.error(f"Error al verificar login: {e}")
            return (False, "Error al verificar credenciales")

    # --- FUNCIONES NUEVAS (FUERA DEL LOGIN) ---

    def obtener_presupuesto_usuario(self, user_id):
        """Busca el presupuesto personalizado de un usuario"""
        try:
            with self.lock:
                self.cursor.execute(f"SELECT presupuesto FROM usuarios WHERE id = {self.p}", (user_id,))
                resultado = self.cursor.fetchone()
                return resultado[0] if resultado else 5000.0
        except Exception as e:
            logging.error(f"Error al obtener presupuesto: {e}")
            return 5000.0

    def actualizar_presupuesto_usuario(self, user_id, nuevo_limite):
        """Actualiza el presupuesto en la tabla de usuarios"""
        try:
            self.cursor.execute(f"UPDATE usuarios SET presupuesto = {self.p} WHERE id = {self.p}", (float(nuevo_limite), user_id))
            self.conexion.commit()
            return True
        except Exception as e:
            logging.error(f"Error al actualizar presupuesto: {e}")
            return False

    def obtener_usuario_por_id(self, user_id):
        """Obtiene los datos de un usuario por su ID"""
        try:
            self.cursor.execute(f"""
                SELECT id, nombre, email, fecha_creacion, presupuesto 
                FROM usuarios 
                WHERE id = {self.p}
            """, (user_id,))
            resultado = self.cursor.fetchone()
            return dict(resultado) if resultado else None
        except Exception as e:
            logging.error(f"Error al obtener usuario: {e}")
            return None


class GestorConPresupuesto(GestorGastos):
    def __init__(self, archivo, limite=5000):
        super().__init__(archivo)
        
        # Tabla de historial de cierres
        self.cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS historial_cierres (
                id {self.serial_type},
                fecha_cierre TEXT,
                total_gastado REAL,
                presupuesto_fijado REAL,
                estado TEXT
            )
        ''')
        self.conexion.commit()
        
        # Limite global de respaldo si no hay usuario
        self.limite_global = limite

    def guardar_presupuesto_global(self, nuevo_limite):
        """Guarda el presupuesto global (obsoleto, pero se mantiene por compatibilidad)"""
        self.limite_global = float(nuevo_limite)
        if self.es_postgresql:
            query = f"INSERT INTO configuracion (clave, valor) VALUES ('presupuesto', {self.p}) ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor"
        else:
            query = f"INSERT OR REPLACE INTO configuracion (clave, valor) VALUES ('presupuesto', {self.p})"
        
        try:
            self.cursor.execute(query, (self.limite_global,))
            self.conexion.commit()
        except:
             # Si no existe la tabla configuración, ignoramos por ahora (es obsoleto)
             pass
    
    def agregar_gasto(self, monto, categoria, descripcion, usuario_id=None, forzar=False):
        # 1. Buscamos el presupuesto REAL del usuario en la tabla usuarios
        limite_usuario = self.obtener_presupuesto_usuario(usuario_id) if usuario_id else self.limite_global
        
        total_actual = self.obtener_total(usuario_id)
        # 2. Validamos contra su presupuesto personal
        if not forzar and (total_actual + float(monto) > limite_usuario):
            return False
        super().agregar_gasto(monto, categoria, descripcion, usuario_id)
        return True

    def consulta_saldo_disponible(self, usuario_id=None):
        gastado = self.obtener_total(usuario_id)
        limite = self.obtener_presupuesto_usuario(usuario_id) if usuario_id else self.limite_global
        return max(0, limite - gastado)
    
    def obtener_porcentaje_usado(self, usuario_id=None):
        gastado = self.obtener_total(usuario_id)
        limite = self.obtener_presupuesto_usuario(usuario_id) if usuario_id else self.limite_global
        if limite == 0: return 0
        return min(100, (gastado / limite) * 100)
    
    def obtener_nivel_alerta(self, usuario_id=None):
        """Retorna el nivel de alerta basado en el porcentaje usado 
        'seguro' (0-79%), 'advertencia' (80-99%), 'peligro' (100%+)
        """
        porcentaje = self.obtener_porcentaje_usado(usuario_id)
        if porcentaje < 80:
            return 'seguro'
        elif porcentaje < 100:
            return 'advertencia'
        else:
            return 'peligro'

    def obtener_categorias_unicas(self):
        """Retorna una lista con todas las categorías únicas registradas en la base de datos"""
        self.cursor.execute("SELECT DISTINCT categoria FROM gastos ORDER BY categoria ASC")
        return [fila[0] for fila in self.cursor.fetchall()]

    
    def obtener_gastos_por_categoria(self, usuario_id=None):
        """Retorna un diccionario con el total gastado por categoría""" 
        with self.lock:
            if usuario_id is not None:
                self.cursor.execute(f"""
                    SELECT categoria, SUM(monto) 
                    FROM gastos 
                    WHERE usuario_id = {self.p}
                    GROUP BY categoria
                """, (usuario_id,))
            else:
                self.cursor.execute("""
                    SELECT categoria, SUM(monto) 
                    FROM gastos 
                    GROUP BY categoria
                """)
            resultados = self.cursor.fetchall()
            return {cat: monto for cat, monto in resultados} 

    def obtener_gastos_por_dia(self, usuario_id=None):
        """Retorna un diccionario con el total gastado por día""" 
        with self.lock:
            if usuario_id is not None:
                self.cursor.execute(f"""
                    SELECT fecha, SUM(monto) 
                    FROM gastos 
                    WHERE usuario_id = {self.p}
                    GROUP BY fecha
                    ORDER BY fecha ASC
                """, (usuario_id,))
            else:
                self.cursor.execute("""
                    SELECT fecha, SUM(monto) 
                    FROM gastos 
                    GROUP BY fecha
                    ORDER BY fecha ASC
                """)
            resultados = self.cursor.fetchall()
            return {fecha: monto for fecha, monto in resultados}

    def obtener_resumen(self, usuario_id):
        """Retorna un resumen financiero completo del usuario encapsulando toda la lógica"""
        try:
            # 1. Obtener todos los gastos de la BD (filtrados por usuario)
            lista_gastos = self.obtener_todos_los_gastos(usuario_id)
            
            # 2. Calcular total general
            total = sum(gasto["monto"] for gasto in lista_gastos)
            
            # 3. Calcular totales por categoría
            desglose = {}
            for gasto in lista_gastos:
                cat = gasto["categoria"]
                monto = gasto["monto"]
                desglose[cat] = desglose.get(cat, 0) + monto
            
            # 4. Obtener información complementaria
            presupuesto = self.obtener_presupuesto_usuario(usuario_id)
            saldo = self.consulta_saldo_disponible(usuario_id=usuario_id)
            porcentaje = self.obtener_porcentaje_usado(usuario_id=usuario_id)
            alerta = self.obtener_nivel_alerta(usuario_id=usuario_id)
            balance = presupuesto - total
            
            # 5. Devolver estructura completa
            return {
                "total_general": total,
                "por_categoria": desglose,
                "presupuesto_limite": presupuesto,
                "saldo_disponible": saldo,
                "porcentaje_usado": porcentaje,
                "nivel_alerta": alerta,
                "balance_neto": balance
            }
        except Exception as e:
            logging.error(f"Error en GestorConPresupuesto.obtener_resumen: {e}")
            raise e

    def datos_para_graficos(self):
        """Retorna los datos para graficos""" 
        self.cursor.execute("SELECT categoria, SUM(monto) FROM gastos GROUP BY categoria")
        resultados = self.cursor.fetchall()
        return {cat: monto for cat, monto in resultados}

    def cerrar_mes(self):
        """Archiva el mes actual en el historial y limpia los gastos detalle"""
        try:
            from datetime import datetime
            fecha_hoy = datetime.now().strftime("%Y-%m-%d %H:%M")
            self.cursor.execute("SELECT SUM(monto) FROM gastos")
            res = self.cursor.fetchone()
            total_actual = res[0] if res and res[0] else 0.0
            presupuesto_actual = self.limite_global
            estado = "Bajo Control" if total_actual <= presupuesto_actual else "Excedido"
            self.cursor.execute(f"""
                INSERT INTO historial_cierres (fecha_cierre, total_gastado, presupuesto_fijado, estado) 
                VALUES ({self.p}, {self.p}, {self.p}, {self.p})
            """, (fecha_hoy, total_actual, presupuesto_actual, estado))
            self.cursor.execute("DELETE FROM gastos")
            self.conexion.commit()
            return True, total_actual
        except Exception as e:
            self.conexion.rollback()
            return False, 0.0


    def obtener_historial_cierres(self):
        """Retorna todos los cierres mensuales guardados como diccionarios"""
        self.cursor.execute("SELECT fecha_cierre, total_gastado, presupuesto_fijado, estado FROM historial_cierres ORDER BY id DESC")
        resultados = self.cursor.fetchall()
        return [dict(fila) for fila in resultados]

    def obtener_gastos_por_mes(self) -> List[Tuple[str, float]]:
        """
        Devuelve una lista de (mes, total_gastado) donde `mes` tiene el formato 'YYYY-MM'.
        """
        self.cursor.execute("""
            SELECT fecha_cierre, total_gastado
            FROM historial_cierres
            ORDER BY fecha_cierre;
        """)
        rows = self.cursor.fetchall()
        
        # Agrupar por mes
        gastos_por_mes: dict[str, float] = {}
        for fecha_str, total in rows:
            # Convertir a datetime y extraer año-mes
            try:
                fecha = datetime.fromisoformat(fecha_str.split()[0])  # Tomar solo la parte de fecha
                mes = fecha.strftime("%Y-%m")
                gastos_por_mes[mes] = gastos_por_mes.get(mes, 0) + total
            except:
                continue
        # Convertir a lista ordenada
        return sorted(gastos_por_mes.items())

    def exportar_a_excel_completo(self, ruta_archivo):
        """
        Exporta gastos actuales e historial de cierres a un archivo Excel con 2 hojas.
        
        Args:
            ruta_archivo (str): Ruta donde guardar el archivo .xlsx
        
        Returns:
            bool: True si se exportó correctamente, False si hubo error
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Crear un nuevo libro de Excel
            wb = Workbook()
            
            # ========== HOJA 1: GASTOS DEL MES ACTUAL ==========
            ws1 = wb.active
            ws1.title = "Gastos Mes Actual"
            
            # Encabezados de la Hoja 1
            encabezados_gastos = ["ID", "Monto", "Categoría", "Fecha", "Descripción"]
            ws1.append(encabezados_gastos)
            
            # Estilo para encabezados (fondo morado, texto blanco, negrita)
            header_fill = PatternFill(start_color="BB86FC", end_color="BB86FC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Obtener datos de gastos actuales
            gastos = self.obtener_todos_los_gastos()
            for gasto in gastos:
                ws1.append(list(gasto.values()))
            
            # Ajustar ancho de columnas
            ws1.column_dimensions['A'].width = 8
            ws1.column_dimensions['B'].width = 12
            ws1.column_dimensions['C'].width = 18
            ws1.column_dimensions['D'].width = 15
            ws1.column_dimensions['E'].width = 40
            
            # ========== HOJA 2: HISTORIAL DE CIERRES ==========
            ws2 = wb.create_sheet(title="Historial Cierres")
            
            # Encabezados de la Hoja 2
            encabezados_historial = ["Fecha Cierre", "Total Gastado", "Presupuesto", "Estado"]
            ws2.append(encabezados_historial)
            
            # Aplicar mismo estilo a encabezados
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Obtener datos del historial de cierres
            historial = self.obtener_historial_cierres()
            for cierre in historial:
                ws2.append(list(cierre.values()))
            
            # Ajustar ancho de columnas
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 15
            ws2.column_dimensions['D'].width = 18
            
            # Guardar el archivo
            wb.save(ruta_archivo)
            return True
        except Exception as e:
            return False


def crear_grafico_gastos_por_mes(datos: List[Tuple[str, float]]):
    """
    `datos` = [(mes, total_gastado), ...]
    Devuelve la figura de Matplotlib lista para incrustar.
    """
    if not datos:
        raise ValueError("No hay datos de cierre para graficar.")

    meses, totales = zip(*datos)  # separa en dos tuplas

    # Paleta premium: tonos azul-gris con gradiente
    if len(meses) == 1:
        colores = [plt.cm.Blues(0.6)]  # Un solo color si solo hay un mes
    else:
        colores = plt.cm.Blues([0.4 + 0.5 * (i / (len(meses)-1)) for i in range(len(meses))])

    fig, ax = plt.subplots(figsize=(8, 4), dpi=100)
    barras = ax.bar(meses, totales, color=colores, edgecolor="#2c3e50")

    # Etiquetas y estilo
    ax.set_title("Gastos mensuales (historial de cierres)", fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Mes", fontsize=12)
    ax.set_ylabel("Total gastado ($)", fontsize=12)
    ax.tick_params(axis='x', rotation=45)
    ax.grid(axis='y', linestyle='--', alpha=0.5)

    # Añadir valores encima de cada barra
    for barra in barras:
        altura = barra.get_height()
        ax.annotate(f"{altura:,.2f}",
                    xy=(barra.get_x() + barra.get_width() / 2, altura),
                    xytext=(0, 4),  # 4 puntos arriba
                    textcoords="offset points",
                    ha='center', va='bottom', fontsize=9, color="#2c3e50")

    fig.tight_layout()
    return fig               